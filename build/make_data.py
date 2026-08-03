#!/usr/bin/env python3
"""원자료(source/*.xlsx) → 도구 데이터 파일(data/*.js) 재생성.

원자료가 갱신되면 source/ 의 파일을 교체하고 이 스크립트를 실행한 뒤,
data/version.js 의 기준일·출처를 함께 수정하세요.

생성 파일
    data/shelters.js   대피장소        ← 화학사고대피장소_목록.xlsx
    data/cases.js      실제 발송 사례   ← 화학사고_재난문자_실제발송사례.xlsx
    data/materials.js  물질정보 460종   ← 화학사고_현장대응_물질정보.hwpx
    data/stats.js      사고통계        ← 화학사고_통계.xls

    pip install openpyxl xlrd
    python3 build/make_data.py
"""
import collections
import json
import pathlib
import re
import sys

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
SRC = ROOT / "source"
OUT = ROOT / "data"

SHELTER_XLSX = SRC / "화학사고대피장소_목록_2026-05-19.xlsx"
CASE_XLSX = SRC / "화학사고_재난문자_실제발송사례.xlsx"
MATERIAL_HWPX = SRC / "화학사고_현장대응_물질정보_460종_2026.hwpx"
STATS_XLS = SRC / "화학사고_통계_2015-2025.xls"
SHELTER_BASE_DATE = "2026-05-19"


def build_shelters() -> int:
    ws = openpyxl.load_workbook(SHELTER_XLSX, data_only=True)["전체 대피장소 목록"]
    tree: dict = collections.OrderedDict()

    for row in ws.iter_rows(min_row=2, values_only=True):
        sido, sgg, kind = row[1], row[2], row[4]
        name, detail, addr, cap, deleted = row[5], row[6], row[7], row[11], row[19]
        if deleted == "Y":
            continue
        addr = (addr or "").strip()
        for prefix in (f"{sido} {sgg} ", f"{sido} "):   # 시도·시군구는 키로 이미 알고 있음
            if addr.startswith(prefix):
                addr = addr[len(prefix):]
                break
        tree.setdefault(sido, collections.OrderedDict()).setdefault(sgg, []).append(
            [name or "", (detail or "").strip(), addr, int(cap or 0), kind or ""]
        )

    for sgg_map in tree.values():
        for items in sgg_map.values():
            items.sort(key=lambda x: x[0])

    total = sum(len(v) for m in tree.values() for v in m.values())
    meta = {
        "기준일자": SHELTER_BASE_DATE,
        "총건수": total,
        "시도수": len(tree),
        "시군구수": sum(len(m) for m in tree.values()),
        "출처": "화학물질안전원 화학사고 대피장소 목록",
        "필드": ["대피장소명", "상세시설명", "도로명주소(시도·시군구 제외)", "수용인원", "시설구분"],
    }

    with open(OUT / "shelters.js", "w", encoding="utf-8") as f:
        f.write(f"/* 화학사고 대피장소 — 원자료: {meta['출처']}(기준일 {SHELTER_BASE_DATE})\n")
        f.write("   구조: SHELTERS[시도][시군구] = [[장소명, 상세시설명, 도로명주소, 수용인원, 시설구분], ...]\n")
        f.write("   ※ 자동 생성 파일입니다. build/make_data.py 로 재생성하세요. */\n")
        f.write("var SHELTER_META = " + json.dumps(meta, ensure_ascii=False) + ";\n")
        f.write("var SHELTERS = " + json.dumps(tree, ensure_ascii=False, separators=(",", ":")) + ";\n")
    return total


CASE_MATERIALS = ["황산", "염산", "암모니아", "불소", "염화수소", "염소", "질산",
             "이산화황", "질소", "유독가스", "화학물질"]

END_WORDS = ["상황종료", "수습", "대응 완료", "대응완료", "조치 완료", "조치완료", "복귀",
             "불검출", "미검출", "안심", "종료", "우려없음", "없음을 알려"]
EVAC_WORDS = ["대피명령", "대피하여", "대피 하시", "대피바람", "대피하시", "으로 대피",
              "대피요망", "신속히 대피"]
INDOOR_WORDS = ["창문", "실내 대기", "실내대피", "외출", "접근"]


def classify(msg: str) -> str:
    if any(w in msg for w in END_WORDS):
        return "상황종료·경과"
    if any(w in msg for w in EVAC_WORDS):
        return "주민소산(대피)"
    if any(w in msg for w in INDOOR_WORDS):
        return "실내대피"
    if "우회" in msg:
        return "차량우회"
    return "기타"


def build_cases() -> int:
    ws = openpyxl.load_workbook(CASE_XLSX, data_only=True)["Sheet1"]
    msgs = [r[0].strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    cases = [{
        "물질": next((m for m in CASE_MATERIALS if m in t), "기타"),
        "단계": classify(t),
        "글자수": len(t),
        "본문": t,
    } for t in msgs]

    with open(OUT / "cases.js", "w", encoding="utf-8") as f:
        f.write("/* 화학사고 재난문자 실제 발송 사례 — 참고용(비승인 자료).\n")
        f.write("   ※ 실제 발송된 원문이며 표준(안)과 다르거나 오탈자가 포함될 수 있습니다. */\n")
        f.write("var CASES = " + json.dumps(cases, ensure_ascii=False, indent=0) + ";\n")
    return len(cases)


def build_materials() -> int:
    """물질정보 hwpx → data/materials.js
    주민 안내에 참고할 수 있는 항목만 남기고, 개인보호구·탐지장비·중화방법 등
    현장대응요원용 정보는 제외한다."""
    import subprocess
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        raw_path = pathlib.Path(tmp) / "raw.json"
        subprocess.run([sys.executable, str(ROOT / "build" / "extract_mat.py"),
                        str(MATERIAL_HWPX), str(raw_path)],
                       check=True, cwd=str(ROOT / "build"), stdout=subprocess.DEVNULL)
        raw = json.loads(raw_path.read_text(encoding="utf-8"))

    def num(v):
        m = re.search(r"(\d+(?:\.\d+)?)", str(v or ""))
        return float(m.group(1)) if m else None

    out = []
    for m in raw:
        names = [x.strip() for x in str(m["물질명"]).split(",") if x.strip()]
        aliases = [x for x in (m.get("유사명") or []) if x] + names[1:]
        d = {
            "n": names[0] if names else m["물질명"],
            "e": str(m.get("영문명", "")).split(",")[0].strip(),
            "c": m.get("cas", ""),
            "a": aliases[:8],
            "s": " / ".join(x for x in [m.get("냄새"), m.get("외관")] if x)
                 or (m.get("성상") or [""])[0],
            "vd": m.get("증기밀도", ""), "sg": m.get("비중", ""),
            "fp": m.get("인화점", ""), "el": m.get("폭발한계", ""),
            "fe": m.get("화재폭발", ""),
            "hz": (m.get("nfpa") or {}).get("건강"),
            "d1": m.get("초기이격거리", ""),      # 누출 · 전 방향
            "d2": m.get("방호활동거리", ""),      # 누출 · 풍하방향 낮/밤
            "d3": m.get("화재대피거리", ""),      # 화재 동반 시 반경
            "sy": (m.get("증상") or {}).get("흡입", ""),
            "e1": (m.get("응급") or {}).get("흡입", ""),
            "e2": (m.get("응급") or {}).get("피부", ""),
            "e3": (m.get("응급") or {}).get("눈", ""),
        }
        vd = num(d["vd"])
        if vd is not None:
            d["vdn"] = vd
        out.append({k: v for k, v in d.items() if v not in ("", None, [])})

    with open(OUT / "materials.js", "w", encoding="utf-8") as f:
        f.write("/* 화학사고 현장대응 물질정보 460종 — 원자료: 화학물질안전원 「화학사고 현장대응 물질정보」(2026)\n")
        f.write("   주민 안내에 참고할 수 있는 항목만 발췌했습니다. 개인보호구·탐지장비·중화방법 등\n")
        f.write("   현장대응요원용 정보는 제외했습니다. build/make_data.py 로 재생성하세요.\n")
        f.write("   n물질명 e영문명 c CAS a유사명 s성상 vd증기밀도 sg비중 fp인화점 el폭발한계\n")
        f.write("   fe화재폭발가능성 hz NFPA건강 d1초기이격거리 d2방호활동거리 d3화재대피거리\n")
        f.write("   sy흡입증상 e1/e2/e3 응급조치(흡입/피부/눈) */\n")
        f.write("var MATERIALS = " + json.dumps(out, ensure_ascii=False, separators=(",", ":")) + ";\n")
    return len(out)


def build_stats() -> dict:
    """사고통계 xls → data/stats.js (사고유형 구분 근거 + 물질 정렬용 빈도)"""
    import xlrd

    wb = xlrd.open_workbook(STATS_XLS)
    sh = wb.sheet_by_name("sheet3")
    hdr = [sh.cell_value(0, c) for c in range(sh.ncols)]
    rows = [{hdr[c]: sh.cell_value(r, c) for c in range(sh.ncols)} for r in range(1, sh.nrows)]

    def base(v):                       # "염산(35%)" → "염산"
        return re.sub(r"\s*[\(（][^)）]*[\)）]", "", str(v)).strip()

    types = collections.Counter(str(r["사고유형"]).strip() for r in rows)
    freq = collections.Counter(base(r["제1사고물질"]) for r in rows if base(r["제1사고물질"]))
    by = collections.defaultdict(collections.Counter)
    for r in rows:
        b = base(r["제1사고물질"])
        if b:
            by[b][str(r["사고유형"]).strip()] += 1

    yr = wb.sheet_by_name("sheet1")
    years = {int(yr.cell_value(r, 0)): int(yr.cell_value(r, 1)) for r in range(1, yr.nrows)}

    data = {
        "기간": f"{min(years)}~{max(years)}", "총건수": len(rows),
        "사고유형": dict(types.most_common()),
        "연도별": years,
        "물질빈도": {k: {"n": v, "t": dict(by[k].most_common())} for k, v in freq.most_common(80)},
    }
    with open(OUT / "stats.js", "w", encoding="utf-8") as f:
        f.write("/* 화학사고 통계 — 원자료: 화학물질안전원 화학사고 발생 현황\n")
        f.write("   용도: 사고유형 4개 구분의 근거 + 물질 입력칸 정렬(자주 발생하는 물질 우선).\n")
        f.write("   ※ 도구는 통계로 어떤 판단도 하지 않습니다. build/make_data.py 로 재생성. */\n")
        f.write("var STATS = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n")
    return data


if __name__ == "__main__":
    for path in (SHELTER_XLSX, CASE_XLSX, MATERIAL_HWPX, STATS_XLS):
        if not path.exists():
            sys.exit(f"원자료를 찾을 수 없습니다: {path}")
    print(f"shelters.js   대피장소 {build_shelters():,}건")
    print(f"cases.js      발송사례 {build_cases()}건")
    print(f"materials.js  물질정보 {build_materials()}종")
    st = build_stats()
    print(f"stats.js      사고 {st['총건수']:,}건 {st['기간']} · 유형 {st['사고유형']}")
    print("완료. data/version.js 의 기준일·출처도 함께 확인하세요.")
