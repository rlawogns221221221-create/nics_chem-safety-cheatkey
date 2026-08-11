#!/usr/bin/env python3
"""방제자원 자료 만들기 — 원자료 6종 → data/resources.js

    python3 build/make_resources.py

── 무엇을 담는가 ────────────────────────────────────────────────
  local   지자체 보유 방재장비   437건  source/방제자원/지자체_보유_방재장비.xlsx
  capsule 봄베 회수용 비상캡슐    19건  붙임2. 봄베 회수용 비상캡슐 보유 업체 현황
  toxgas  독성가스 회수업체        2건  2. (사업장) 전국 독성가스 처리업체
  waste   폐기물 수집·운반/처분   19건  DB_260629 … 사업장 연락처
  marine  해경 방제비축기지         3건  (해양경찰) 방제비축기지
  water   수질오염방제센터          4건  수질오염방제센터 방제물품 현황

── 개인정보 ────────────────────────────────────────────────────
원자료에는 담당자 이름·직급·개인 휴대전화·이메일이 들어 있습니다.
**하나도 싣지 않습니다.** 사업장·기관 대표번호(02·031·041… 로 시작하는
유선번호)만 남기고, 010 으로 시작하는 번호는 기계적으로 걸러냅니다.
마지막에 검사해서 개인번호가 하나라도 남으면 만들다 멈춥니다.

── 좌표 ────────────────────────────────────────────────────────
원자료에 좌표가 없어 주소로부터 잡습니다(build/geo_anchor.py). 인터넷
주소검색을 쓸 수 없는 자리라 정확도가 세 등급으로 나뉘며, 그 등급을
자료에 실어 화면에서 구분해 표시합니다. 정확한 좌표가 필요하면
build/geocode.html 을 인터넷 되는 PC 에서 한 번 돌리세요.
"""
import json
import pathlib
import re
import sys

ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "build"))

from geo_anchor import Anchor            # noqa: E402
import res_sources as RS                 # noqa: E402

SRC = ROOT / "source" / "방제자원"
OUT = ROOT / "data" / "resources.js"

# 개인 휴대전화 — 어떤 칸에서든 이 모양이면 버립니다
MOBILE = re.compile(r"01[016789][-\s.]?\d{3,4}[-\s.]?\d{4}")
TELOK = re.compile(r"^0(?:2|[3-6]\d|50\d|70)[-\s.]?\d{3,4}[-\s.]?\d{4}$")


def clean_tel(v):
    """사업장 유선번호만 통과. 개인 휴대전화·빈칸은 ''."""
    t = re.sub(r"\s+", "", str(v or "")).replace("(", "").replace(")", "")
    t = t.split("/")[0].split(",")[0].strip()
    if not t or t == "-":
        return ""
    if MOBILE.fullmatch(t.replace("-", "")) or t.replace("-", "").startswith("01"):
        return ""
    return t if TELOK.match(t) else ""


def norm_sido(v):
    s = re.sub(r"\s+", "", str(v or ""))
    fix = {"울산공역시": "울산광역시", "강원도": "강원특별자치도",
           "전라북도": "전북특별자치도", "전북": "전북특별자치도",
           "충북": "충청북도", "충남": "충청남도", "전남": "전라남도",
           "경북": "경상북도", "경남": "경상남도", "세종시": "세종특별자치시"}
    return fix.get(s, s)


def norm_sgg(v):
    s = re.sub(r"\s+", " ", str(v or "")).strip()
    return "" if s in ("-", "null", "None") else s


def sido_of(addr):
    """주소 앞머리에서 시·도를 뽑는다."""
    a = str(addr or "").strip()
    for full, short in [
        ("서울특별시", "서울"), ("부산광역시", "부산"), ("대구광역시", "대구"),
        ("인천광역시", "인천"), ("광주광역시", "광주"), ("대전광역시", "대전"),
        ("울산광역시", "울산"), ("세종특별자치시", "세종"), ("경기도", "경기"),
        ("강원특별자치도", "강원"), ("충청북도", "충북"), ("충청남도", "충남"),
        ("전북특별자치도", "전북"), ("전라남도", "전남"), ("경상북도", "경북"),
        ("경상남도", "경남"), ("제주특별자치도", "제주"),
    ]:
        if a.startswith(full) or a.startswith(short):
            return full
    return ""


def sgg_of(addr, sido):
    a = str(addr or "")
    a = re.sub(r"^\S+\s*", "", a) if sido else a
    m = re.match(r"\s*([가-힣]+(?:시|군|구))\s+([가-힣]+구)\b", a)
    if m:
        return m.group(1) + " " + m.group(2)
    m = re.match(r"\s*([가-힣]+(?:시|군|구))\b", a)
    return m.group(1) if m else ""


def strip_sgg(addr, sido, sgg):
    """주소에서 시·도/시·군·구를 떼어 낸다 — 목록에 두 번 나오지 않게."""
    a = str(addr or "").strip()
    for p in (sido, sido[:2] if sido else "", sgg):
        if p and a.startswith(p):
            a = a[len(p):].strip()
    return a


# ── 지자체 보유 방재장비 ────────────────────────────────────────
COLS = [
    ("빈 탱크로리", "대"), ("굴착기", "대"), ("암롤차", "대"),
    ("스키드로더", "대"), ("진공흡입차량", "대"), ("지게차", "대"),
    ("방독면", "개"), ("보호복", "개"), ("장갑", "개"), ("장화", "개"),
    ("소석회", "kg"), ("마른모래", "kg"), ("중탄산나트륨", "kg"),
    ("흡착포", ""), ("유처리제", "kg"),
]
# 화면의 '필요한 것으로 찾기' 갈래
GROUP = {
    "빈 탱크로리": "car", "굴착기": "car", "암롤차": "car", "스키드로더": "car",
    "진공흡입차량": "car", "지게차": "car",
    "방독면": "gear", "보호복": "gear", "장갑": "gear", "장화": "gear",
    "소석회": "chem", "마른모래": "chem", "중탄산나트륨": "chem",
    "흡착포": "chem", "유처리제": "chem",
}


def qty(v):
    """'1', '11390', '69박스', '-' → (숫자, 원표기) / 없으면 None."""
    if v is None:
        return None
    s = str(v).replace(",", "").strip()
    if s in ("-", "", "0"):
        return None
    m = re.match(r"^([\d.]+)\s*(.*)$", s)
    if not m:
        return None
    try:
        n = float(m.group(1))
    except ValueError:
        return None
    return (n, m.group(2).strip()) if n > 0 else None


def read_local(anchor):
    import openpyxl
    path = SRC / "지자체_보유_방재장비.xlsx"
    if not path.exists():
        sys.exit(f"원자료가 없습니다: {path}")
    ws = openpyxl.load_workbook(path, data_only=True).worksheets[0]

    out = []
    for r in range(3, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 28)]
        if v[0] is None and not any(v[1:6]):
            continue
        sido = norm_sido(v[2])
        sgg = norm_sgg(v[3])
        dept = re.sub(r"\s+", " ", str(v[4] or "")).strip()
        addr = re.sub(r"\s+", " ", str(v[6] or "")).strip()
        if not (sido and dept):
            continue

        items, groups = [], set()
        for i, (nm, unit) in enumerate(COLS):
            q = qty(v[12 + i])
            if not q:
                continue
            n, tail = q
            num = f"{int(n):,}" if n == int(n) else f"{n:g}"
            items.append(f"{nm} {num}{tail or unit}")
            groups.add(GROUP[nm])
        if not items:
            continue                       # 가진 장비가 없는 부서는 지도에 둘 이유가 없다

        # 시·군·구 칸은 믿기 어렵다 — 시·도 이름이 그대로 들어 있거나
        # 기관 이름이 들어 있어, 주소에서 다시 확정한다
        sgg_key = anchor.sgg_name(sido, "" if sgg == sido else sgg, addr)
        la, lo, acc = anchor.find(sido, sgg_key, addr)
        out.append({
            "t": "local",
            "n": f"{sgg_key or sido} {dept}",
            "org": sgg_key or sido, "dept": dept,
            "a": strip_sgg(addr, sido, sgg_key),
            "sd": sido, "sg": sgg_key or sido,
            "la": la, "lo": lo, "ap": acc,
            "c": " · ".join(items),
            "g": sorted(groups),
            "tel": clean_tel(v[9]),
            "keep": re.sub(r"\s+", " ", str(v[11] or "")).strip(),
            "src": "지자체 보유 방재장비 현황",
        })
    return out


# ── 나머지 다섯 갈래 ────────────────────────────────────────────
def read_capsule(anchor):
    out = []
    for name, addr, spec, tel, note in RS.ERCV:
        sido = sido_of(addr)
        sgg = anchor.sgg_name(sido, sgg_of(addr, sido), addr)
        la, lo, acc = anchor.find(sido, sgg, addr)
        out.append({
            "t": "capsule", "n": name, "a": strip_sgg(addr, sido, sgg),
            "sd": sido, "sg": sgg, "la": la, "lo": lo, "ap": acc,
            "c": spec + (" · " + note if note else ""),
            "g": ["capsule"], "tel": clean_tel(tel),
            "src": "봄베 회수용 비상캡슐 보유 업체 현황",
        })
    return out


def read_toxgas(anchor):
    out = []
    for r in RS.TOXGAS:
        sgg = anchor.sgg_name(r["sido"], r["sgg"], r["addr"])
        la, lo, acc = anchor.find(r["sido"], sgg, r["addr"])
        bits = [r["job"], "처리가능 " + r["gas"], r["car"], "운영인력 " + r["crew"]]
        if r["note"]:
            bits.append(r["note"])
        out.append({
            "t": "toxgas", "n": r["name"],
            "a": strip_sgg(r["addr"], r["sido"], r["sgg"]),
            "sd": r["sido"], "sg": sgg or r["sgg"], "la": la, "lo": lo, "ap": acc,
            "c": " · ".join(bits), "g": ["gas"], "tel": clean_tel(r["tel"]),
            "src": "전국 독성가스 회수업체 현황",
        })
    return out


def read_waste(anchor):
    out = []
    for name, step, addr, tel in RS.WASTE:
        if not addr:
            continue                       # 주소가 없으면 지도에 찍을 수 없다
        sido = sido_of(addr)
        sgg = anchor.sgg_name(sido, sgg_of(addr, sido), addr)
        la, lo, acc = anchor.find(sido, sgg, addr)
        out.append({
            "t": "waste", "n": name, "a": strip_sgg(addr, sido, sgg),
            "sd": sido, "sg": sgg, "la": la, "lo": lo, "ap": acc,
            "c": step, "g": ["waste"], "tel": clean_tel(tel),
            "src": "화학사고 방제자원 통합 구축 사업장 연락처",
        })
    return out


def read_marine():
    return [{
        "t": "marine", "n": r["name"], "a": r["where"],
        "sd": r["sido"], "sg": r["sgg"], "la": r["lat"], "lo": r["lon"],
        "ap": r["acc"], "c": " · ".join(r["stock"]), "g": ["stock"],
        "tel": "", "agency": r["agency"], "note": r["note"],
        "src": "해양경찰 방제비축기지 현황",
    } for r in RS.MARINE]


def read_water():
    return [{
        "t": "water", "n": r["name"], "a": r["where"],
        "sd": r["sido"], "sg": r["sgg"], "la": r["lat"], "lo": r["lon"],
        "ap": r["acc"], "c": " · ".join(r["stock"]), "g": ["stock"],
        "tel": "", "agency": r["agency"], "note": r["note"],
        "src": "수질오염방제센터 방제물품 현황",
    } for r in RS.WATER]


KINDS = [
    {"id": "local", "이름": "지자체 보유 장비",
     "쓰임": "관내 시·군·구가 직접 가진 차량·보호구·중화제. 협조 요청이 가장 빠릅니다"},
    {"id": "capsule", "이름": "비상캡슐(ERCV)",
     "쓰임": "누출된 가스 봄베를 통째로 담아 밀봉·이송합니다"},
    {"id": "toxgas", "이름": "독성가스 회수업체",
     "쓰임": "암모니아·프레온 등 독성가스를 회수·처리합니다"},
    {"id": "waste", "이름": "폐기물 처리업체",
     "쓰임": "회수한 오염물·폐기물을 수집·운반하고 처분합니다"},
    {"id": "marine", "이름": "해경 방제비축기지",
     "쓰임": "해양오염 방제 물자를 대량으로 비축한 곳입니다"},
    {"id": "water", "이름": "수질오염방제센터",
     "쓰임": "하천·호소로 흘러든 오염물질을 막고 걷어냅니다"},
]

# '무엇이 필요한가'로 찾기 — 종류를 몰라도 필요한 물건으로 찾게
NEEDS = [
    {"id": "car", "이름": "차량·중장비", "설명": "탱크로리·굴착기·암롤차·진공흡입차량·지게차"},
    {"id": "gear", "이름": "보호구", "설명": "방독면·보호복·장갑·장화"},
    {"id": "chem", "이름": "중화제·흡착재", "설명": "소석회·마른모래·중탄산나트륨·흡착포·유처리제"},
    {"id": "capsule", "이름": "봄베 밀봉", "설명": "누출 용기를 담아 옮길 비상캡슐"},
    {"id": "gas", "이름": "가스 회수", "설명": "독성가스 회수·처리"},
    {"id": "waste", "이름": "폐기물 처리", "설명": "오염물 수집·운반·소각·매립"},
    {"id": "stock", "이름": "대량 비축", "설명": "오일펜스·흡착재 등을 대량으로 가진 국가 비축기지"},
]

ACC_NOTE = {
    "road": "같은 도로명에 있는 다른 시설 좌표로 잡음 (오차 수백 m)",
    "emd": "같은 읍·면·동 평균 좌표 (오차 1~3km)",
    "sgg": "시·군·구 대표점 (오차 시·군 크기)",
    "sido": "시·도 대표점 — 본청 부서라 시·군·구가 없음 (오차 큼)",
    "port": "항만 위치",
    "base": "유역 거점 위치",
    "exact": "주소검색으로 확인한 좌표",
}


def main():
    anchor = Anchor()
    rows = (read_local(anchor) + read_capsule(anchor) + read_toxgas(anchor)
            + read_waste(anchor) + read_marine() + read_water())
    rows = [r for r in rows if r["la"] is not None]
    rows.sort(key=lambda r: (r["sd"], r["sg"], r["t"], r["n"]))

    # ── 개인정보 검사 — 하나라도 남으면 만들다 멈춘다 ──
    blob = json.dumps(rows, ensure_ascii=False)
    hit = MOBILE.findall(blob)
    if hit:
        sys.exit(f"개인 휴대전화가 남아 있습니다: {hit[:5]}")
    for bad in ("이메일", "@"):
        if bad in blob:
            sys.exit(f"'{bad}' 가 자료에 남아 있습니다")

    meta = {
        "기준일": "2026-08-11",
        "총건수": len(rows),
        "담당자포함": False,
        "좌표": "주소로 잡은 어림값 — 정확도는 각 건의 ap 값 참고",
        "출처": [k["이름"] for k in KINDS],
    }
    head = (
        "/* 화학사고 방제자원 — 공개용\n"
        "   담당자 개인 이름·직급·휴대전화·이메일은 들어 있지 않습니다.\n"
        "   사업장·기관 대표번호(유선)만 싣습니다.\n"
        "\n"
        "   구조: RESOURCES = [{t, n, a, sd, sg, la, lo, ap, c, g, tel, src}, ...]\n"
        "     t   종류 — local 지자체장비 / capsule 비상캡슐 / toxgas 독성가스\n"
        "                waste 폐기물 / marine 해경비축기지 / water 수질방제센터\n"
        "     n   이름            a  주소(시·도·시군구 제외)\n"
        "     sd  시도            sg 시군구\n"
        "     la  위도            lo 경도\n"
        "     ap  좌표 정확도 — road/emd/sgg/port/base/exact (RESOURCE_ACC 참고)\n"
        "     c   보유·처리 내용 (원표기 그대로 · 검색 대상)\n"
        "     g   무엇에 쓰나 (RESOURCE_NEEDS 의 id 들)\n"
        "     tel 사업장·기관 대표번호   src 출처 자료명\n"
        "     agency 운영기관 (marine·water — 대표번호가 원자료에 없음)\n"
        "\n"
        "   ※ 자동 생성 파일입니다. build/make_resources.py 로 재생성하세요. */\n"
    )
    body = [head]
    body.append("var RESOURCE_META = " + json.dumps(meta, ensure_ascii=False) + ";")
    body.append("var RESOURCE_KINDS = " + json.dumps(KINDS, ensure_ascii=False) + ";")
    body.append("var RESOURCE_NEEDS = " + json.dumps(NEEDS, ensure_ascii=False) + ";")
    body.append("var RESOURCE_ACC = " + json.dumps(ACC_NOTE, ensure_ascii=False) + ";")
    body.append("var RESOURCES = [")
    body.append(",\n".join(json.dumps(r, ensure_ascii=False, separators=(",", ":"))
                           for r in rows))
    body.append("];")
    OUT.write_text("\n".join(body) + "\n", encoding="utf-8")

    kinds = {}
    for r in rows:
        kinds[r["t"]] = kinds.get(r["t"], 0) + 1
    print(f"{OUT.relative_to(ROOT)}  {len(rows)}건  "
          f"{OUT.stat().st_size / 1024:.0f}KB")
    for k in KINDS:
        print(f"   {k['이름']:<16} {kinds.get(k['id'], 0):>4}건")
    print("   좌표 정확도:", anchor.stat)
    notel = sum(1 for r in rows if not r["tel"] and not r.get("agency"))
    print(f"   대표번호 없음 {notel}건 — 화면에 '번호 확인 필요'로 표시됩니다")
    print("   개인정보 검사 통과 — 이름·휴대전화·이메일 없음")


if __name__ == "__main__":
    main()
